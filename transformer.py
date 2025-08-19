import pandas as pd
import re  # 正则表达式模块
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def parse_questions(text):
    # 预处理文本
    text = text.replace("&#xA;", "\n").replace("&lt;", "<").replace("&gt;", ">")

    # 分割题目
    question_blocks = re.split(r'\n(?=\d+\.\d+\.\d+\.\s+第\d+题)', text)
    questions = []

    for block in question_blocks:
        if not block.strip():
            continue

        # 提取题号
        qid_match = re.search(r'(\d+\.\d+\.\d+)\.\s+第(\d+)题', block)
        if not qid_match:
            continue

        full_qid = qid_match.group(1)
        qnum = qid_match.group(2)

        # 提取题型和等级
        type_level = full_qid.split('.')
        q_type = {
            '1': '单选题',
            '2': '多选题',
            '3': '判断题'
        }.get(type_level[0], '未知题型')

        # 只处理选择题和判断题
        if q_type not in ['单选题', '多选题', '判断题']:
            continue

        q_level = {
            '1': '初级工', '2': '中级工', '3': '高级工',
            '4': '技师', '5': '高级技师'
        }.get(type_level[1], '未知等级')

        # 提取题干
        content_start = block.find('\n', qid_match.end()) + 1
        content_end = re.search(r'\n(?:[A-D]\.|正确答案)', block[content_start:])
        if content_end:
            content = block[content_start:content_start + content_end.start()].strip()
        else:
            content = block[content_start:].split('正确答案：')[0].strip()

        # 清理题干内容
        # 去除"第x页"字样及其变体
        content = re.sub(r'第\s*\d+\s*页\s*[:：]?', '', content).strip()
        # 去除多余的空格和换行
        content = re.sub(r'\s+', ' ', content).strip()

        # 提取选项（仅选择题）
        options = {}
        if q_type in ['单选题', '多选题']:
            for opt in ['A', 'B', 'C', 'D']:
                opt_match = re.search(rf'{opt}\.(.*?)(?:\n|$)', block)
                if opt_match:
                    options[opt] = opt_match.group(1).strip()

        # 提取答案
        answer_match = re.search(r'正确答案[:：]\s*([A-D]+|[正确错误]+)', block)
        answer = answer_match.group(1) if answer_match else ""

        # 提取评价点
        eval_match = re.search(r'关联评价点的名称[:：]\s*(.+)', block)
        eval_point = eval_match.group(1).strip() if eval_match else ""

        questions.append({
            "题型": q_type,
            "等级": q_level,
            "题号": full_qid,
            "题目编号": int(qnum),
            "题目内容": content,
            "选项A": options.get('A', ''),
            "选项B": options.get('B', ''),
            "选项C": options.get('C', ''),
            "选项D": options.get('D', ''),
            "正确答案": answer,
            "关联评价点": eval_point,
            "工种": "变电设备检修工(开关)",
            "工种定义": "从事电网变电站一次开关类设备验收、维护、检修的人员"
        })

    return questions


def create_excel(questions):
    df = pd.DataFrame(questions)

    # 优化列顺序
    df = df[['题型', '等级', '题号', '题目编号', '题目内容',
             '选项A', '选项B', '选项C', '选项D',
             '正确答案', '关联评价点', '工种', '工种定义']]

    # 创建Excel
    with pd.ExcelWriter('变电设备检修工(开关)技能题库.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='题库')

        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['题库']

        # 设置列宽
        col_widths = {
            'A': 8,  # 题型
            'B': 10,  # 等级
            'C': 10,  # 题号
            'D': 8,  # 题目编号
            'E': 80,  # 题目内容
            'F': 30,  # 选项A
            'G': 30,  # 选项B
            'H': 30,  # 选项C
            'I': 30,  # 选项D
            'J': 8,  # 正确答案
            'K': 35,  # 关联评价点
            'L': 15,  # 工种
            'M': 50  # 工种定义
        }

        for col, width in col_widths.items():
            col_letter = get_column_letter(ord(col) - 64)
            worksheet.column_dimensions[col_letter].width = width

        # 设置自动换行
        wrap_columns = ['E', 'F', 'G', 'H', 'I', 'K', 'M']
        for col in wrap_columns:
            col_letter = get_column_letter(ord(col) - 64)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.alignment = Alignment(wrapText=True, vertical='top')

        # 添加筛选器
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    return 'Excel文件已生成：变电设备检修工(开关)技能题库.xlsx'


# ==== 主程序开始 ====
if __name__ == "__main__":
    # 1. 从文件读取题库文本
    try:
        with open('题库文本.txt', 'r', encoding='utf-8') as f:
            text_content = f.read()
        print("题库文本读取成功！")
    except FileNotFoundError:
        print("错误：找不到题库文本.txt文件！")
        print("请确保文件与脚本在同一目录下。")
        exit(1)
    except UnicodeDecodeError:
        print("错误：文件编码问题！尝试使用GBK编码...")
        try:
            with open('题库文本.txt', 'r', encoding='gbk') as f:
                text_content = f.read()
            print("使用GBK编码读取成功！")
        except:
            print("错误：无法读取文件！请检查文件编码。")
            exit(1)

    # 2. 解析题目
    print("开始解析题目...")
    questions = parse_questions(text_content)
    print(f"成功解析 {len(questions)} 道题目！")

    # 3. 生成Excel
    print("开始生成Excel文件...")
    result = create_excel(questions)
    print(result)